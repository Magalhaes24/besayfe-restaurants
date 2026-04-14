from __future__ import annotations

import io

from openpyxl import load_workbook


async def extract_xlsx(file_bytes: bytes) -> dict:
    """
    Read all sheets from XLSX. Return structured representation of all cells.
    Handles merged cells by expanding them.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
    # data_only=True: get computed cell values (not formulas)

    sheets_data: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Unmerge and fill merged cells
        _expand_merged_cells(ws)

        rows = []
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [_serialize_cell(c) for c in row]
            # Skip entirely empty rows
            if any(c is not None for c in cells):
                if headers is None:
                    headers = cells  # First non-empty row as headers
                else:
                    rows.append(dict(zip(headers, cells)))

        sheets_data[sheet_name] = {
            "headers": headers or [],
            "rows": rows,
            "row_count": len(rows),
        }

    # Build raw text representation for Claude
    raw_text = _sheets_to_text(sheets_data)

    return {
        "sheets": sheets_data,
        "sheet_names": wb.sheetnames,
        "raw_text": raw_text,
    }


def _expand_merged_cells(ws) -> None:
    """Fill merged cell ranges with the top-left cell value."""
    for merged_range in list(ws.merged_cells.ranges):
        min_row, min_col = merged_range.min_row, merged_range.min_col
        top_left_value = ws.cell(min_row, min_col).value
        ws.unmerge_cells(str(merged_range))
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=merged_range.max_row,
            min_col=min_col,
            max_col=merged_range.max_col,
        ):
            for cell in row:
                cell.value = top_left_value


def _serialize_cell(value) -> str | None:
    """Convert cell value to string; handle None, float formatting."""
    if value is None:
        return None
    if isinstance(value, float):
        # Avoid scientific notation for small numbers
        return f"{value:g}"
    return str(value)


def _sheets_to_text(sheets_data: dict) -> str:
    parts = []
    for sheet_name, data in sheets_data.items():
        parts.append(f"=== Sheet: {sheet_name} ===")
        headers = data.get("headers") or []
        parts.append(" | ".join(str(h) for h in headers))
        for row in data.get("rows", []):
            parts.append(" | ".join(str(v) for v in row.values()))
    return "\n".join(parts)
