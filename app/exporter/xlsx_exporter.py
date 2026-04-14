from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.models.schema import NormalizedMenuSheet

# Color palette (dark navy header, light alternating rows)
HEADER_BG = "1F4E79"
HEADER_FG = "FFFFFF"
ALT_ROW_BG = "DCE6F1"
ACCENT_BG = "2E75B6"
TOTAL_BG = "FFF2CC"


def _header_font() -> Font:
    return Font(bold=True, color=HEADER_FG, size=11)


def _header_fill() -> PatternFill:
    return PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def _thin_border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def export_to_xlsx(sheet: NormalizedMenuSheet) -> bytes:
    """
    Export NormalizedMenuSheet to a formatted XLSX workbook.
    Returns bytes for streaming in the HTTP response.
    """
    wb = Workbook()

    # Sheet 1: Recipe summary
    ws_summary = wb.active
    ws_summary.title = "Ficha Técnica"
    _write_summary(ws_summary, sheet)

    # Sheet 2: Ingredients
    ws_ingredients = wb.create_sheet("Ingredientes")
    _write_ingredients(ws_ingredients, sheet)

    # Sheet 3: Steps
    ws_steps = wb.create_sheet("Preparação")
    _write_steps(ws_steps, sheet)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_summary(ws, sheet: NormalizedMenuSheet) -> None:
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 35

    # Title
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = sheet.name
    title_cell.font = Font(bold=True, size=14, color=ACCENT_BG)
    title_cell.alignment = _center()
    ws.row_dimensions[1].height = 28

    fields = [
        ("Categoria", sheet.category),
        ("País", sheet.country),
        ("Região", sheet.region),
        ("Continente", sheet.continent),
        ("Porções", sheet.servings),
        ("Tempo de Preparação (min)", sheet.prep_time_minutes),
        ("Tempo de Confeção (min)", sheet.cooking_time_minutes),
        ("Custo Total (€)", sheet.total_cost),
        ("Custo por Porção (€)", sheet.cost_per_serving),
        ("Confiança IA", f"{sheet.confidence_score:.0%}"),
        ("Ficheiro Fonte", sheet.source_file),
        ("Formato Fonte", sheet.source_format.upper()),
        ("Normalizado em", sheet.normalized_at.strftime("%Y-%m-%d %H:%M UTC")),
    ]

    for i, (label, value) in enumerate(fields, start=2):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.font = Font(bold=True)
        label_cell.fill = PatternFill(
            start_color="DEEAF1", end_color="DEEAF1", fill_type="solid"
        )
        label_cell.border = _thin_border()

        value_cell = ws.cell(row=i, column=2, value=value)
        value_cell.border = _thin_border()

        # Format currency cells
        if "€" in label:
            value_cell.number_format = '#,##0.00 €'


def _write_ingredients(ws, sheet: NormalizedMenuSheet) -> None:
    headers = [
        "Nº",
        "Produto",
        "Quantidade",
        "Unidade",
        "Preço Unitário (€)",
        "Custo (€)",
        "Observações",
    ]
    col_widths = [6, 28, 14, 10, 20, 14, 30]

    for i, (header, width) in enumerate(zip(headers, col_widths), start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 22
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = _center()
        cell.border = _thin_border()

    for row_idx, ing in enumerate(sheet.ingredients, start=2):
        is_alt = row_idx % 2 == 0
        row_data = [
            ing.line_number,
            ing.product_name,
            ing.quantity,
            ing.unit,
            ing.unit_price,
            ing.line_cost,
            ing.observations or "",
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _thin_border()
            if is_alt:
                cell.fill = PatternFill(
                    start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid"
                )
            # Format currency and number columns
            if col_idx in (3, 5, 6):
                cell.number_format = '#,##0.0000'
            if col_idx in (5, 6):
                cell.number_format = '#,##0.00 €'

    # Total row
    total_row = len(sheet.ingredients) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=6, value=sheet.total_cost)
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00 €'
    total_cell.fill = PatternFill(
        start_color=TOTAL_BG, end_color=TOTAL_BG, fill_type="solid"
    )

    per_serving_cell = ws.cell(row=total_row + 1, column=6, value=sheet.cost_per_serving)
    ws.cell(
        row=total_row + 1, column=1, value=f"Por porção ({sheet.servings} porções)"
    ).font = Font(bold=True)
    per_serving_cell.font = Font(bold=True)
    per_serving_cell.number_format = '#,##0.00 €'
    per_serving_cell.fill = PatternFill(
        start_color=TOTAL_BG, end_color=TOTAL_BG, fill_type="solid"
    )


def _write_steps(ws, sheet: NormalizedMenuSheet) -> None:
    headers = ["Nº", "Ação", "Produto", "Tipo Corte", "Tempo (min)", "Observações"]
    col_widths = [6, 20, 25, 15, 12, 35]

    for i, (header, width) in enumerate(zip(headers, col_widths), start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 22
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = _center()
        cell.border = _thin_border()

    for row_idx, step in enumerate(sheet.steps, start=2):
        is_alt = row_idx % 2 == 0
        row_data = [
            step.step_number,
            step.action,
            step.product or "",
            step.cut_type or "",
            step.time_minutes,
            step.observations or "",
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _thin_border()
            if is_alt:
                cell.fill = PatternFill(
                    start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid"
                )
